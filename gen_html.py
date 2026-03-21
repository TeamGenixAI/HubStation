#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_html.py
엑셀 파일(ParmForSpring.xlsx, DoForSpring.xlsx, DiForSpring.xlsx, AiForSpring.xlsx)로부터
testmerge.html 자동 생성

규칙:
  DoForSpring.xlsx  → 라디오 버튼 박스
  ParmForSpring.xlsx
    OnMsg == "bitbit"  → 체크박스 박스
    그 외              → number input 박스
  DiForSpring.xlsx  → 상태 탭: label(id=DI{InstanceNo}), 초기값=OffMsg
  AiForSpring.xlsx  → 상태 탭: label(id=AI{InstanceNo}), 초기값=TagName+OffMsg

  TagCategory → 탭 이름
  GroupName   → tagBoxDisplay 박스의 ribbon 텍스트
  OnMsg/OffMsg (DO) → 첫 번째/두 번째 라디오 라벨
  TagName (bitbit)  → 체크박스 for-label 내용
  TagName (number)  → number input 앞 레이블
  UnitName (number) → number input 뒤 단위

빈 행 처리:
  동일 GroupName 사이 빈 행 → tagBoxDisplay 내부 여백 (width:3rem)
  다른 GroupName 사이 빈 행 → 다음 tagBoxDisplay를 새 ROW에서 시작
"""

import math
import openpyxl
from collections import OrderedDict

# ─────────────────────────────────────────────
# 엑셀 읽기
# ─────────────────────────────────────────────
def read_xlsx(path):
    """DO용: 빈 행 무시"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    result = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            result.append(dict(zip(headers, row)))
    return result


def read_xlsx_with_spacers(path):
    """Parm용: 빈 행을 {'_spacer': True}로 보존"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            result.append({'_spacer': True})
        elif any(v is not None for v in row):
            result.append(dict(zip(headers, row)))
    return result


def preprocess_spacers(rows):
    """
    빈 행(_spacer)을 분류하고 TagCategory 상속:
      전후 GroupName이 같으면 → _spacer_type='inner'    (tagBoxDisplay 내 여백)
      전후 GroupName이 다르면 → _spacer_type='row_break' (새 row 시작)
    """
    result = []
    for i, row in enumerate(rows):
        if row.get('_spacer'):
            prev = next((rows[j] for j in range(i - 1, -1, -1) if not rows[j].get('_spacer')), None)
            nxt  = next((rows[j] for j in range(i + 1, len(rows)) if not rows[j].get('_spacer')), None)
            prev_grp = prev.get('GroupName') if prev else None
            next_grp = nxt.get('GroupName')  if nxt  else None
            spacer_type = 'inner' if (prev_grp and next_grp and prev_grp == next_grp) else 'row_break'
            result.append({
                '_spacer': True,
                '_spacer_type': spacer_type,
                'TagCategory': prev.get('TagCategory') if prev else None,
            })
        else:
            result.append(row)
    return result


# ─────────────────────────────────────────────
# 개별 요소 HTML 조각 생성
# ─────────────────────────────────────────────
def radio_inner_html(item):
    """라디오 버튼 HTML (ribbon 제외)
    UnitName == MODE : id = DO{n}  / DO{n+1}           (2개)
    UnitName == DOW  : id = DO{n} / DO{n+1} / DO{n}S  (3개: OffMsg / OnMsg / 정지)
    그 외            : id = DO{n}T / DO{n}F            (2개)
    """
    n         = item['InstanceNo']
    on_label  = item.get('OnMsg')  or '켜짐'
    off_label = item.get('OffMsg') or '꺼짐'
    unit      = (item.get('UnitName') or 'NORMAL').upper()

    radio_div = (
        '<div class="form-check" style="color:rgb(0,0,150);font-size:2rem;'
        'position:relative;display:inline-flex;{margin}width:auto;">'
        '<input class="form-check-input" type="radio" id="{rid}" name="DO{n}G" '
        'value="{val}" onclick="DoActionClick(this)">'
        '<label class="form-check-label" for="{rid}">{label}</label></div>'
    )

    def make_radio(rid, val, label, margin=''):
        return radio_div.format(n=n, rid=rid, val=val, label=label,
                                margin=f'margin-left:{margin};' if margin else '')

    if unit == 'MODE':
        return (
            make_radio(f'DO{n}',     'schedule', on_label) +
            make_radio(f'DO{n + 1}', 'manual',   off_label, '0.7rem')
        )
    elif unit == 'DOW':
        return (
            make_radio(f'DO{n}',     'CLOSE', off_label) +
            make_radio(f'DO{n + 1}', 'OPEN',  on_label,  '0.7rem') +
            make_radio(f'DO{n}S',    'PAUSE', '정지',     '0.7rem')
        )
    else:  # NORMAL
        return (
            make_radio(f'DO{n}T', 'start', on_label) +
            make_radio(f'DO{n}F', 'stop',  off_label, '0.7rem')
        )


def checkbox_inner_html(item):
    """체크박스 + 라벨 HTML (ribbon 제외)"""
    uid      = f"P{item['InstanceNo']}B{item.get('OffMsg', '')}"
    tag_name = item.get('TagName') or ''

    return (
        f'<div class="form-check" style="color:rgb(0,0,150);font-size:2rem;'
        f'position:relative;display:inline-flex;width:auto;margin-right:0.5rem;">'
        f'<input class="form-check-input" type="checkbox" id="{uid}" '
        f'style="font-weight:bold;border-width:2px;border-style:solid;">'
        f'<label class="form-check-label" for="{uid}" '
        f'style="margin:2px;"><small style="margin:0.2rem;">{tag_name}</small></label></div>'
    )


def number_inner_html(item):
    """TagName + number input + UnitName HTML (ribbon 제외)"""
    n        = item['InstanceNo']
    tag_name = item.get('TagName') or ''
    unit     = item.get('UnitName') or ''

    label_html = f'<small style="margin:0.2rem;">{tag_name}</small>' if tag_name else ''
    unit_html  = f'<small style="margin:3px;">{unit}</small>'        if unit else ''

    return (
        f'<div class="d-inline-flex align-items-center" style="margin:0.2rem;">'
        f'{label_html}'
        f'<input type="number" inputmode="decimal" id="P{n}" style="width:4.5rem;">'
        f'{unit_html}'
        f'</div>'
    )


# ─────────────────────────────────────────────
# 상태 탭 전용: DI / AI 표시 요소 HTML
# ─────────────────────────────────────────────
def di_inner_html(item):
    """DI 상태 label HTML (ribbon 제외)
    ID = DI{InstanceNo}, 초기 내용 = OffMsg만 표시
    """
    n       = item['InstanceNo']
    off_msg = str(item.get('OffMsg') or '')

    return (
        f'<small id="DI{n}" style="margin:0.6rem;font-size:2rem;">{off_msg}</small>'
    )


def ai_inner_html(item):
    """AI 상태 label HTML (ribbon 제외)
    ID = A{InstanceNo}, 초기 내용 = OffMsg만
    TagName은 앞에 <small>, UnitName은 뒤에 <small>로 표시
    """
    n        = item['InstanceNo']
    tag_name = str(item.get('TagName') or '')
    off_msg  = str(item.get('OffMsg')  or '')
    unit     = str(item.get('UnitName') or '')

    tag_html  = f'<small style="margin:0.2rem;">{tag_name}</small>' if tag_name else ''
    unit_html = f'<small style="margin:0.2rem;">{unit}</small>'     if unit     else ''

    return (
        f'<div class="d-inline-flex align-items-center" style="margin:0.2rem;">'
        f'{tag_html}'
        f'<label id="A{n}" style="font-size:1.8rem;font-weight:bold;">{off_msg}</label>'
        f'{unit_html}'
        f'</div>'
    )


def make_status_group_box(group_name, items, source):
    """DI/AI 항목 그룹 → tagBoxDisplay (상태 탭 전용)
    source: 'di' | 'ai'
    ribbon 텍스트 길이 기반으로 min-width 설정 → 박스가 항상 ribbon보다 넓음
    """
    ribbon = f'<span class="ribbon">{group_name}</span>'
    # 한글 1자 ≈ 1.3rem, 영문/숫자 ≈ 0.8rem, 여유 +1rem
    min_w = sum(1.3 if ord(c) > 127 else 0.8 for c in group_name) + 1.0
    if source == 'di':
        inner = ''.join(di_inner_html(it) for it in items)
    else:
        inner = ''.join(ai_inner_html(it) for it in items)
    return (
        '<div class="d-flex align-items-center flex-wrap tagBoxDisplay" '
        f'style="width:auto;padding-left:10px;padding-right:0px;min-width:{min_w:.1f}rem;">'
        f'{inner}{ribbon}</div>'
    )


def build_status_boxes(di_items, ai_items):
    """DI + AI 항목 → 상태 탭용 박스 리스트"""
    boxes = []

    # DI: GroupName별로 묶기
    di_groups = OrderedDict()
    for it in di_items:
        gn = it.get('GroupName') or ''
        di_groups.setdefault(gn, []).append(it)
    for gn, group_items in di_groups.items():
        boxes.append((make_status_group_box(gn, group_items, 'di'), False, False))

    # AI: GroupName별로 묶기
    # '실내습도'는 row break 추가 → 양액기/교반기/에어포그 가동시간 row와 분리
    ROW_BREAK_BEFORE = {'실내습도'}
    ai_groups = OrderedDict()
    for it in ai_items:
        gn = it.get('GroupName') or ''
        ai_groups.setdefault(gn, []).append(it)
    for gn, group_items in ai_groups.items():
        row_break = gn in ROW_BREAK_BEFORE
        boxes.append((make_status_group_box(gn, group_items, 'ai'), row_break, False))

    return boxes


# ─────────────────────────────────────────────
# GroupName 그룹 → 하나의 tagBoxDisplay 박스
# ─────────────────────────────────────────────
def make_group_box(group_name, items, source):
    """
    같은 GroupName의 항목들을 하나의 tagBoxDisplay로 합친다.
    source: 'do' | 'parm'
    items에 {'_inner_spacer': True}가 포함될 수 있음 → 내부 여백 div 삽입
    """
    ribbon = f'<span class="ribbon">{group_name}</span>'

    if source == 'do':
        inner = ''.join(radio_inner_html(it) for it in items)
        has_dow = any((it.get('UnitName') or '').upper() == 'DOW' for it in items)
        wrap_cls = 'flex-nowrap' if has_dow else 'flex-wrap'
        style    = 'width:auto;' if has_dow else 'max-width:15rem;'
        html = (
            f'<div class="d-xxl-flex justify-content-between {wrap_cls} tagBoxDisplay" '
            f'style="{style}">'
            f'{inner}{ribbon}</div>'
        )
        return html, False
    else:
        # 빈 행(_inner_spacer) 기준으로 그룹 분리
        groups = []
        current = []
        for it in items:
            if it.get('_inner_spacer'):
                if current:
                    groups.append(current)
                    current = []
            else:
                if str(it.get('OnMsg', '')).lower() == 'bitbit':
                    current.append(checkbox_inner_html(it))
                else:
                    current.append(number_inner_html(it))
        if current:
            groups.append(current)

        if len(groups) > 1:
            # 빈 행으로 구분된 그룹 2개 이상 → 각 그룹을 inline-flex로 묶고 space-between
            # → width:100% 단독 row 차지
            group_html = ''.join(
                f'<div class="d-inline-flex align-items-center">{"".join(g)}</div>'
                for g in groups
            )
            html = (
                '<div class="d-flex align-items-center flex-wrap tagBoxDisplay full-row" '
                'style="width:100%;padding-left:10px;padding-right:0px;justify-content:space-between;">'
                f'{group_html}{ribbon}</div>'
            )
            return html, True   # full_row=True → 단독 row
        else:
            inner = ''.join(groups[0]) if groups else ''
            html = (
                '<div class="d-flex align-items-center flex-wrap tagBoxDisplay" '
                'style="width:auto;padding-left:10px;padding-right:0px;">'
                f'{inner}{ribbon}</div>'
            )
            return html, False


# ─────────────────────────────────────────────
# 탭 레이아웃 생성
# ─────────────────────────────────────────────
def build_tab_boxes(do_items, parm_items):
    """
    한 탭에 속한 DO/Parm 항목을 GroupName으로 묶어 tagBoxDisplay 박스 리스트로 변환.
    parm_items에 spacer dict가 포함될 수 있음.
    반환: list of (html_str, row_break_before)
    """
    boxes = []  # (html, row_break_before, full_row)

    # DO: GroupName별로 묶기 (순서 유지)
    do_groups = OrderedDict()
    for it in do_items:
        gn = it.get('GroupName') or ''
        do_groups.setdefault(gn, []).append(it)
    for gn, group_items in do_groups.items():
        box_html, is_full = make_group_box(gn, group_items, 'do')
        boxes.append((box_html, is_full, is_full))

    # Parm: 순서대로 처리 (spacer 포함)
    current_group = None
    current_items = []
    pending_row_break = False

    for it in parm_items:
        if it.get('_spacer'):
            if it['_spacer_type'] == 'inner':
                current_items.append({'_inner_spacer': True})
            else:  # row_break
                if current_group is not None:
                    box_html, is_full = make_group_box(current_group, current_items, 'parm')
                    boxes.append((box_html, pending_row_break or is_full, is_full))
                    pending_row_break = False
                    current_group = None
                    current_items = []
                pending_row_break = True
        else:
            gn = it.get('GroupName') or ''
            if gn != current_group:
                if current_group is not None:
                    box_html, is_full = make_group_box(current_group, current_items, 'parm')
                    boxes.append((box_html, pending_row_break or is_full, is_full))
                    pending_row_break = False
                current_group = gn
                current_items = []
            current_items.append(it)

    if current_group is not None:
        box_html, is_full = make_group_box(current_group, current_items, 'parm')
        boxes.append((box_html, pending_row_break or is_full, is_full))

    return boxes


def boxes_to_rows_html(boxes, indent='                ', min_rows=3, min_per_row=2):
    """
    박스 리스트 → d-flex flex-row 행 HTML
    boxes: list of (html_str, row_break_before, full_row)
    - row_break_before=True : 해당 박스 앞에서 강제 줄바꿈
    - full_row=True         : 단독 row 차지 (앞뒤 모두 줄바꿈, width:100% 박스)
    - 나머지는 기존 알고리즘(items_per_row) 적용
    """
    if not boxes:
        return ''

    n = len(boxes)
    items_per_row = max(min_per_row, math.ceil(n / min_rows))

    html = ''
    current_row = []

    def flush_row():
        nonlocal html, current_row
        if not current_row:
            return
        inner = ('\n' + indent + '    ').join(current_row)
        html += (
            f'{indent}<div class="d-flex flex-row justify-content-between '
            f'align-items-center flex-wrap" style="width:100%;">\n'
            f'{indent}    {inner}\n'
            f'{indent}</div>\n'
        )
        current_row = []

    for box_html, row_break, full_row in boxes:
        if (row_break or full_row) and current_row:
            flush_row()
        current_row.append(box_html)
        if len(current_row) >= items_per_row or full_row:
            flush_row()

    flush_row()
    return html


# ─────────────────────────────────────────────
# 탭 네비게이션 + 탭 패널 HTML 생성
# ─────────────────────────────────────────────
def build_tabs_html(tab_order, do_by_tab, parm_by_tab, status_boxes=None, status_bottom_html=''):
    nav_items = []
    pane_items = []

    for idx, tab_name in enumerate(tab_order):
        tab_id   = f'tab-{idx + 1}'
        is_first = (idx == 0)
        active   = 'active' if is_first else ''

        nav_items.append(
            f'        <li class="nav-item" role="presentation">'
            f'<a class="nav-link {active}" role="tab" data-bs-toggle="tab" '
            f'href="#{tab_id}" style="font-size:16px;font-weight:bold;">'
            f'{tab_name}</a></li>'
        )

        if tab_name == '상태' and status_boxes is not None:
            boxes = status_boxes
        else:
            do_items   = do_by_tab.get(tab_name, [])
            parm_items = parm_by_tab.get(tab_name, [])
            boxes      = build_tab_boxes(do_items, parm_items)
        # 시간 탭: 박스가 넓어 1280px에서 4개 배치 시 wrapping → 3개/행으로 강제
        rows_html = boxes_to_rows_html(boxes, min_rows=4 if tab_name == '시간' else 3)
        if tab_name == '상태' and status_bottom_html:
            rows_html += status_bottom_html

        pane_items.append(
            f'            <div class="tab-pane {active}" role="tabpanel" '
            f'id="{tab_id}" style="width:100%;height:auto;">\n'
            f'                <div class="container d-flex flex-column justify-content-between flex-wrap" '
            f'style="padding-left:1.5rem;padding-right:1.5rem;border-style:none;padding-top:0.5rem;margin:0.3rem auto;">\n'
            f'{rows_html}'
            f'                </div>\n'
            f'            </div>'
        )

    nav_html  = '\n'.join(nav_items)
    pane_html = '\n'.join(pane_items)
    return nav_html, pane_html


# ─────────────────────────────────────────────
# 전체 HTML 조립
# ─────────────────────────────────────────────
HTML_TEMPLATE = """\
<!DOCTYPE html>
<html data-bs-theme="light" lang="ko" style="height:90vh;font-size:1rem;">

<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, shrink-to-fit=no">
    <title>PTY3 - Auto Generated</title>
    <link rel="stylesheet" href="assets/bootstrap/css/bootstrap.min.css">
    <link rel="stylesheet" href="assets/css/bss-overrides.css">
    <link rel="stylesheet" href="assets/css/MainPannel.css">
    <link rel="stylesheet" href="assets/css/styles.css">
    <link rel="stylesheet" href="assets/css/merged-layout.css">
</head>

<body class="w-100">
    <div style="width:100%;height:auto;">
        <div style="width:100%;height:auto;">
            <ul class="nav nav-tabs d-flex flex-row" role="tablist"
                style="margin:5px;padding:1px;width:auto;height:auto;">
{NAV_ITEMS}
            </ul>
            <div class="tab-content" style="width:100%;height:auto;">
{PANE_ITEMS}
            </div>
        </div>
    </div>

      <div class="toast fade text-center bg-info position-fixed top-50 start-50 translate-middle hide" role="alert"
        id="toast-1" data-bs-delay="2500" style="margin:auto;font-size:30px;">
        <div class="toast-body" role="alert" style="font-size:30px;">
            <p class="fs-6 fw-semibold" id="ToastMsg"
                style="font-size:xx-large;color:var(--bs-body-bg);text-align:center;">Toast message</p>
        </div>
    </div>
    <script>
        let lastTap = 0;

        function toggleFullscreen() {
            if (!document.fullscreenElement) {
                document.documentElement.requestFullscreen();
            } else {
                document.exitFullscreen();
            }
        }

        document.addEventListener('touchend', function (e) {
            // 터치 이벤트가 실행된 시간
            const now = new Date().getTime();
            if (now - lastTap < 300) { // 300ms 이내에 두 번 터치
                toggleFullscreen();
            }
            lastTap = now;
        });

        // PC 환경(마우스 더블클릭)도 지원하고 싶다면 아래 추가
        document.addEventListener('dblclick', toggleFullscreen);
    </script>
    <script>
        document.querySelectorAll('input[type="radio"], input[type="checkbox"]').forEach(function (el) {
            el.addEventListener('focus', function (e) {
                e.target.blur();
            });
        });

    </script>

    <script type="text/javascript" th:inline="javascript">
        let sockwtInfo = [[${ wsuri }]];
        let aiinform = [[${ aiInform }]];
        let parmvalues = [[${ parmValues }]];
        let parminform = [[${ parmInform }]];
        let doinform = [[${ doInform }]];
        let diinform = [[${ diInform }]];
    </script>

    <script src="assets/bootstrap/js/bootstrap.min.js"></script>
    <script src="assets/js/jquery-3.7.1.min.js"></script>
    <script src="js/merged-websocket.js"></script>
</body>
</html>
"""


def main():
    print("엑셀 파일 읽는 중...")
    do_data   = read_xlsx('DoForSpring.xlsx')
    parm_raw  = read_xlsx_with_spacers('ParmForSpring.xlsx')
    parm_data = preprocess_spacers(parm_raw)
    di_data   = read_xlsx('DiForSpring.xlsx')
    ai_data   = read_xlsx('AiForSpring.xlsx')

    # ── 탭 순서 수집 (등장 순서 유지, 중복 제거) ──
    tab_order = list(OrderedDict.fromkeys(
        [r['TagCategory'] for r in do_data   if r.get('TagCategory')] +
        [r['TagCategory'] for r in parm_data if r.get('TagCategory') and not r.get('_spacer')]
    ))

    # 상태 탭 추가 (DI/AI 모두 TagCategory='상태')
    if '상태' not in tab_order:
        tab_order.append('상태')

    # ── 탭별 데이터 분류 ──
    do_by_tab   = {}
    parm_by_tab = {}

    for row in do_data:
        cat = row.get('TagCategory')
        if cat:
            do_by_tab.setdefault(cat, []).append(row)

    for row in parm_data:
        cat = row.get('TagCategory')
        if cat:
            parm_by_tab.setdefault(cat, []).append(row)

    # ── 상태 탭 박스 생성 (현재시각 제외 → 하단 고정 행으로 분리) ──
    jigak_items  = [r for r in ai_data if r.get('GroupName') == '현재시각']
    ai_data_main = [r for r in ai_data if r.get('GroupName') != '현재시각']
    status_boxes = build_status_boxes(di_data, ai_data_main)

    # 하단 고정 행: 현재시각(좌) + BtnDownload(우)
    jigak_box = make_status_group_box('현재시각', jigak_items, 'ai') if jigak_items else ''
    ind = '                '
    status_bottom_html = (
        f'{ind}<div class="d-flex flex-row justify-content-between align-items-center flex-wrap" style="width:100%;">\n'
        f'{ind}    {jigak_box}\n'
        f'{ind}    <div class="tagDisplay110" style="width:auto;">'
        f'<button class="btn btn-primary" id="BtnDownload" style="font-size:30px;" '
        f'type="button" onclick="btnDownload()">제어기로 설정치 쓰기</button></div>\n'
        f'{ind}</div>\n'
    )

    # ── HTML 생성 ──
    nav_html, pane_html = build_tabs_html(tab_order, do_by_tab, parm_by_tab, status_boxes, status_bottom_html)

    html = HTML_TEMPLATE.replace('{NAV_ITEMS}',  nav_html)
    html = html.replace('{PANE_ITEMS}', pane_html)

    out_path = 'testmerge.html'
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"완료: {out_path}")
    print(f"  탭 수       : {len(tab_order)} ({', '.join(tab_order)})")
    print(f"  DO 그룹     : {len(set(r.get('GroupName','') for r in do_data))}개")
    parm_groups = set(r.get('GroupName','') for r in parm_data if not r.get('_spacer'))
    print(f"  Parm 그룹   : {len(parm_groups)}개")
    print(f"  DI 그룹     : {len(set(r.get('GroupName','') for r in di_data))}개 ({len(di_data)}항목)")
    print(f"  AI 그룹     : {len(set(r.get('GroupName','') for r in ai_data))}개 ({len(ai_data)}항목)")
    print(f"  전체 아이템 : {sum(1 for r in parm_data if not r.get('_spacer'))}개")
    print(f"  행 구분자   : {sum(1 for r in parm_data if r.get('_spacer_type') == 'row_break')}개")


if __name__ == '__main__':
    main()
