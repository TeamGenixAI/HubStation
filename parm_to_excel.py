"""
ParmForSpring.json → Excel (xlsx) 변환 스크립트
실행: python parm_to_excel.py
결과: ParmForSpring.xlsx (같은 폴더에 생성)
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "ParmForSpring.json"
OUTPUT_FILE = "ParmForSpring.xlsx"

COLUMNS = [
    ("InstanceNo", 12),
    ("TagCategory", 13),
    ("GroupName", 16),
    ("TagName", 30),
    ("UnitName", 10),
    ("OnMsg", 10),
    ("OffMsg", 10),
    ("Max", 10),
    ("Min", 10),
    ("Step", 10),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def main():
    with open(INPUT_FILE, encoding="utf-8") as f:
        data = json.load(f)

    items = data.get("informs", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ParmForSpring"

    # 헤더 행
    for col_idx, (col_key, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, item in enumerate(items, start=2):
        for col_idx, (col_key, _) in enumerate(COLUMNS, start=1):
            value = item.get(col_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 필터 추가
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    print(f"저장 완료: {OUTPUT_FILE}  ({len(items)}개 항목)")

if __name__ == "__main__":
    main()
